import openpyxl
from openpyxl.styles import Font

def write_student_to_excel(student_data, template_path, output_path):
    """
    Writes student responses and scores into the CEP Excel sheet.
    - student_data: dict with name, id, responses, subject_scores
    - template_path: path to the Excel template
    - output_path: path to save the updated sheet
    """
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Class10"]

    # Find next empty row
    row = sheet.max_row + 1

    sheet.cell(row=row, column=1).value = row - 1  # S.No.
    sheet.cell(row=row, column=2).value = student_data["name"]
    sheet.cell(row=row, column=3).value = student_data["student_id"]
    sheet.cell(row=row, column=4).value = student_data.get("section", "A")

    # Fill responses (Q1–Q108)
    for q_no in range(1, 109):
        col = 5 + q_no - 1
        value = 1 if student_data["responses"].get(q_no) == student_data["answer_key"].get(q_no) else 0
        sheet.cell(row=row, column=col).value = value

    # Optional: Highlight correct answers
    for q_no in student_data["correct"]:
        col = 5 + q_no - 1
        sheet.cell(row=row, column=col).font = Font(bold=True)

    wb.save(output_path)
    print(f"📤 Saved results to {output_path}")
